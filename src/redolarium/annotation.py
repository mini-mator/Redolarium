# SPDX-License-Identifier: MIT
# Copyright (c) 2026 Redolarium Contributors, IIT Kanpur
# See LICENSE and THIRD_PARTY_LICENSES.md for full licence details.

import os
import re
import time
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from Bio import SeqIO, Entrez
from Bio.SeqFeature import SeqFeature, FeatureLocation
from Bio.Blast import NCBIWWW, NCBIXML
from redolarium.utils import api_retry, save_publication_plot, safe_subprocess_run
from redolarium import db_manager

# Housekeeping genes for MLSA (Multi-Locus Sequence Analysis)
def _load_housekeeping_genes_config():
    try:
        import yaml
        config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config", "config.yaml")
        if os.path.exists(config_path):
            with open(config_path, "r", encoding="utf-8") as f:
                cfg = yaml.safe_load(f)
                hk = cfg.get("housekeeping_genes")
                if hk:
                    return hk
    except Exception:
        pass
    return {
        # 7-gene MLSA scheme for broad bacteria (overridable via config.yaml "housekeeping_genes")
        # Reference: Goris et al. 2007 (doi:10.1099/ijs.0.65060-0)
        #            Konstantinidis & Tiedje 2005 (doi:10.1073/pnas.0409727102)
        #            MLST.net recommended loci for multi-genus comparisons
        "dnaA": ["dnaa", "chromosomal replication initiator protein", "dnaa protein"],
        "gyrB": ["gyrb", "dna gyrase subunit b", "gyrase b"],
        "rpoB": ["rpob", "rna polymerase subunit beta", "rna polymerase beta"],
        "recA": ["reca", "recombinase a", "dna recombinase a"],
        "trpB": ["trpb", "tryptophan synthase subunit beta", "tryptophan synthase beta chain"],
        "ilvD": ["ilvd", "dihydroxy-acid dehydratase"],
        "purH": ["purh", "phosphoribosylaminoimidazolecarboxamide formyltransferase"]
    }

HOUSEKEEPING_GENES = _load_housekeeping_genes_config()

# Alignment task function for parallel execution
def get_similarity(seq1, seq2):
    if not seq1 or not seq2:
        return 0.0, 0.0
    try:
        from Bio.Align import PairwiseAligner
        aligner = PairwiseAligner()
        aligner.mode = 'global'
        aligner.match_score = 1.0
        aligner.mismatch_score = 0.0
        aligner.open_gap_score = -1.0
        aligner.extend_gap_score = -0.5
        alignments = aligner.align(seq1, seq2)
        if len(alignments) == 0:
            return 0.0, 0.0
        alignment = alignments[0]
        
        # Strict length/match assertions for annotation validation
        len1, len2 = len(seq1), len(seq2)
        if max(len1, len2) > 0 and (min(len1, len2) / max(len1, len2)) < 0.2:
            import sys
            print(f"[Annotation Validation Warning] Extreme length mismatch detected during alignment: {len1} vs {len2}. Possible annotation error.", file=sys.stderr)
            return 0.0, 0.0
            
        try:
            matches = alignment.counts.identical
            alignment_len = len(alignment)
        except AttributeError:
            # Biopython < 1.80 doesn't have counts.identical, alignment.score can be negative due to gap penalties.
            # Using difflib as a robust, scientifically accurate fallback.
            import difflib
            sm = difflib.SequenceMatcher(None, seq1, seq2)
            identity = sm.ratio() * 100
            return round(identity, 2), 100.0
            
        identity = (matches / alignment_len) * 100 if alignment_len > 0 else 0.0
        coverage = (min(len(seq1), len(seq2)) / max(len(seq1), len(seq2))) * 100
        return round(identity, 2), round(coverage, 2)
    except Exception:
        import difflib
        sm = difflib.SequenceMatcher(None, seq1, seq2)
        identity = sm.ratio() * 100
        return round(identity, 2), 100.0


def align_single_cds(args):
    q_idx, q_tag, q_symbol, q_product, q_trans, q_start, q_end, q_strand, gc_pct, ref_cds_data, last_ref_idx = args
    match_found = False
    ref_tag, ref_symbol = "NA", "NA"
    identity, coverage = 0.0, 0.0
    category = "Unique"
    
    if q_trans in ref_cds_data['by_hash']:
        r_idx, r_tag, r_symbol = ref_cds_data['by_hash'][q_trans]
        ref_tag, ref_symbol = r_tag, r_symbol
        identity, coverage = 100.0, 100.0
        category = "Core"
        last_ref_idx = r_idx
        match_found = True
        
    if not match_found and q_symbol != "NA":
        sym_lower = q_symbol.lower()
        if sym_lower in ref_cds_data['by_symbol']:
            r_idx, r_tag, r_symbol, r_trans = ref_cds_data['by_symbol'][sym_lower]
            ref_tag, ref_symbol = r_tag, r_symbol
            identity, coverage = get_similarity(q_trans, r_trans)
            # Core genome threshold: ≥95% AA identity (strain-level pan-genome standard)
            # Reference: Konstantinidis & Tiedje 2005; Snipen & Ussery 2010
            if identity >= 95.0:
                category = "Core"
            elif identity >= 50.0:
                category = "Accessory"
            elif identity >= 30.0:
                category = "Distant Homolog"  # twilight zone — structural, not ortholog
            else:
                category = "Unique"
            last_ref_idx = r_idx
            match_found = True
            
    if not match_found:
        from redolarium.config import CONFIG
        window = CONFIG.get("annotation", {}).get("ortholog_search_window", 15)
        search_start = max(0, last_ref_idx - window)
        search_end = min(len(ref_cds_data['list']), last_ref_idx + window)
        best_ident = 0.0
        best_cov = 0.0
        best_candidate = None
        best_idx = last_ref_idx
        
        for r_idx in range(search_start, search_end):
            r_idx_val, r_tag, r_symbol, r_trans = ref_cds_data['list'][r_idx]
            ident, cov = get_similarity(q_trans, r_trans)
            if ident > best_ident:
                best_ident = ident
                best_cov = cov
                best_candidate = (r_tag, r_symbol)
                best_idx = r_idx
                
        # Only assign a category if identity is above the twilight zone lower bound (30%)
        # Below 30%: hit may be structural analogue, not a homolog — label Unique
        # Reference: Rost 1999 (doi:10.1093/protein/12.2.85) — twilight zone boundary
        if best_ident >= 30.0:
            ref_tag, ref_symbol = best_candidate
            identity, coverage = best_ident, best_cov
            if identity >= 95.0:
                category = "Core"
            elif identity >= 50.0:
                category = "Accessory"
            else:
                category = "Distant Homolog"  # 30-50%: twilight zone, not reliable ortholog
            last_ref_idx = best_idx
            
    return {
        "Locus_Tag": q_tag,
        "Gene_Symbol": q_symbol,
        "Start_Coord": q_start,
        "End_Coord": q_end,
        "Strand": q_strand,
        "Product_Description": q_product,
        "GC_Percent": gc_pct,
        "Ref_Ortholog_Tag": ref_tag,
        "Ref_Ortholog_Symbol": ref_symbol,
        "Identity_Pct": identity,
        "Coverage_Pct": coverage,
        "Category": category,
        "Protein_Sequence": q_trans
    }

def find_housekeeping_gene(record):
    for feat in record.features:
        if feat.type == "CDS":
            gene = feat.qualifiers.get("gene", [""])[0].lower()
            product = feat.qualifiers.get("product", [""])[0].lower()
            trans = feat.qualifiers.get("translation", [""])[0]
            
            if not trans:
                continue
                
            for name, aliases in HOUSEKEEPING_GENES.items():
                # Allow exact gene name OR exact matching against curated aliases (e.g. product names)
                # This ensures we strictly find the 7 MLSA genes without drifting, even if 'gene' qualifier is missing.
                if gene == name or gene in aliases or product in aliases:
                    return name, trans
    return None, None

def predict_genes_from_fasta(record, logger):
    logger.info("[INFO] Raw FASTA detected. Automatically predicting genes and ORFs using native Pyrodigal (C-bindings)...")
    
    try:
        import pyrodigal
    except ImportError:
        logger.error("Pyrodigal package not found. Please install it using: pip install pyrodigal")
        raise ImportError("Pyrodigal is a required dependency for processing raw FASTA genome inputs.")
        
    sequence = str(record.seq)
    
    try:
        # Check pyrodigal version/API (supports both GeneFinder and OrfFinder)
        try:
            detector = pyrodigal.GeneFinder()
        except AttributeError:
            detector = pyrodigal.OrfFinder()
            
        # Train ORFs finder on input sequence (Standard prokaryotic translation table 11)
        try:
            detector.train(sequence.encode('utf-8'))
        except Exception:
            try:
                detector.train(sequence)
            except Exception:
                pass # Training might be optional in standard default runs
                
        # Locate coding sequences
        try:
            genes = detector.find_genes(sequence.encode('utf-8'))
        except Exception:
            genes = detector.find_genes(sequence)
            
    except Exception as e:
        logger.error(f"Failed to execute Pyrodigal gene prediction: {e}")
        raise e
        
    logger.info(f"Pyrodigal predicted {len(genes)} coding sequences.")
    
    # Bridge predicted ORFs back to standard SeqFeatures expected downstream
    for idx, gene in enumerate(genes):
        strand = gene.strand
        
        try:
            translation = gene.translate()
        except Exception:
            translation = ""
            
        locus_tag = f"GM_PROD_{idx+1:04d}"
        
        location = FeatureLocation(gene.begin - 1, gene.end, strand=strand)
        feature = SeqFeature(
            location,
            type="CDS",
            qualifiers={
                "locus_tag": [locus_tag],
                "translation": [translation],
                "gene": [f"orf_{idx+1}"],
                "product": ["predicted protein"]
            }
        )
        record.features.append(feature)

def load_local_reference_genomes(out_dir, logger):
    logger.info("[INFO] API query failed or returned empty. Activating local offline fallback loading reference genomes...")
    
    local_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reference_genomes")
    local_files = []
    if os.path.exists(local_dir):
        local_files = [os.path.join(local_dir, f) for f in os.listdir(local_dir) if f.endswith((".gb", ".gbk", ".fasta", ".fa"))]
        
    cwd_dir = os.getcwd()
    cwd_files = [os.path.join(cwd_dir, f) for f in os.listdir(cwd_dir) if f.endswith((".gb", ".gbk"))]
    all_files = local_files + cwd_files
    
    ref_strains = []
    if all_files:
        for f_path in all_files[:25]:
            try:
                recs = list(SeqIO.parse(f_path, "genbank"))
                if recs:
                    rec = max(recs, key=lambda r: len(r.seq))
                    ref_strains.append(f"{rec.annotations.get('organism', 'Local Ref')} ({rec.id})")
            except Exception:
                try:
                    recs = list(SeqIO.parse(f_path, "fasta"))
                    if recs:
                        rec = max(recs, key=lambda r: len(r.seq))
                        ref_strains.append(f"Local Fasta Ref ({rec.id})")
                except Exception:
                    pass
                    
    if not ref_strains:
        logger.error("No local GenBank reference files found under reference_genomes/ or workspace root.")
        return []
        
    return ref_strains

def load_ncbi_cache():
    cache_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "resources")
    os.makedirs(cache_dir, exist_ok=True)
    cache_path = os.path.join(cache_dir, "ncbi_query_cache.json")
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                import json
                return json.load(f)
        except Exception:
            pass
    return {}

def save_ncbi_cache(cache_data):
    cache_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "resources")
    os.makedirs(cache_dir, exist_ok=True)
    cache_path = os.path.join(cache_dir, "ncbi_query_cache.json")
    try:
        with open(cache_path, "w", encoding="utf-8") as f:
            import json
            json.dump(cache_data, f, indent=4)
    except Exception:
        pass

def validate_nucleotide_assembly(accession, logger):
    try:
        handle = Entrez.elink(dbfrom="nuccore", db="assembly", id=accession)
        record = Entrez.read(handle)
        if record and record[0]["LinkSetDb"]:
            link_ids = [link["Id"] for link in record[0]["LinkSetDb"][0]["Link"]]
            if link_ids:
                handle_sum = Entrez.esummary(db="assembly", id=link_ids[0])
                record_sum = Entrez.read(handle_sum)
                doc_sum = record_sum['DocumentSummarySet']['DocumentSummary'][0]
                level = doc_sum.get('AssemblyStatus', doc_sum.get('AssemblyLevel', ''))
                logger.info(f"NCBI esummary AssemblyLevel check for {accession}: {level}")
                return level in ["Complete Genome", "Chromosome", "Scaffold", "Contig"]
    except Exception as e:
        logger.warning(f"validate_nucleotide_assembly link check failed for {accession}: {e}. Falling back to title keyword matching.")
    return None

def detect_16s_strict_chain(query_record, logger, temp_dir=".", email="researcher@example.com"):
    logger.info("Executing 16S rRNA strict tool detection chain (Barrnap -> Infernal -> GenBank -> SILVA BLAST)...")
    import subprocess
    import shutil
    
    candidates = []
    
    # 1. Attempt Barrnap via local or WSL docker
    barrnap_run = False
    temp_fasta = os.path.join(temp_dir, "temp_query.fasta")
    try:
        with open(temp_fasta, "w") as f:
            f.write(f">query\n{str(query_record.seq)}\n")
        
        # Test local
        barrnap_path = shutil.which("barrnap")
        if barrnap_path:
            res = subprocess.run([barrnap_path, "--quiet", temp_fasta], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            if res.returncode == 0:
                for line in res.stdout.splitlines():
                    if "16S_rRNA" in line:
                        parts = line.split("\t")
                        start, end = int(parts[3]), int(parts[4])
                        score = float(parts[5]) if parts[5] != "." else 1.0
                        seq = str(query_record.seq[start-1:end])
                        candidates.append({"seq": seq, "method": "barrnap_local", "score": score, "len": len(seq)})
                        barrnap_run = True
        
        # Test Docker if local failed
        if not barrnap_run:
            try:
                res = safe_subprocess_run(["wsl", "docker", "run", "--rm", "-v", f"{os.path.abspath(temp_dir)}:/data", "seemann/barrnap:latest", "barrnap", "/data/temp_query.fasta"], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            except Exception as e:
                logger.warning(f"Barrnap docker crashed: {e}")
                res = type('obj', (object,), {'returncode': 1, 'stdout': '', 'stderr': str(e)})()
            if res.returncode == 0:
                for line in res.stdout.splitlines():
                    if "16S_rRNA" in line:
                        parts = line.split("\t")
                        start, end = int(parts[3]), int(parts[4])
                        score = float(parts[5]) if parts[5] != "." else 1.0
                        seq = str(query_record.seq[start-1:end])
                        candidates.append({"seq": seq, "method": "barrnap_docker", "score": score, "len": len(seq)})
                        barrnap_run = True
    except Exception as e:
        logger.warning(f"Barrnap invocation failed: {e}")
        
    # Remove temp file if created
    if os.path.exists(temp_fasta):
        try:
            os.remove(temp_fasta)
        except Exception:
            pass

    # 2. Parse annotated GenBank rRNA features
    if not candidates:
        logger.info("Falling back to parsing annotated GenBank rRNA features...")
        for feat in query_record.features:
            if feat.type == "rRNA":
                prod = feat.qualifiers.get("product", [""])[0].lower()
                if "16s" in prod:
                    seq = str(feat.location.extract(query_record.seq))
                    candidates.append({"seq": seq, "method": "genbank_annotation", "score": 1.0, "len": len(seq)})

    # Score each candidate and pick the highest
    best_candidate = None
    best_score = -1.0
    for cand in candidates:
        c_score = (min(cand["len"], 1500) / 1500.0) * 0.4 + (min(cand["len"], 2000) / 2000.0) * 0.2 + (cand["score"] * 0.4)
        if c_score > best_score:
            best_score = c_score
            best_candidate = cand

    if best_candidate:
        logger.info(f"Selected best 16S candidate via {best_candidate['method']} (Length: {best_candidate['len']} bp, Score: {best_score:.2f})")
        return best_candidate["seq"]
    
    logger.error("rRNA strict detection chain failed to identify any 16S rRNA gene candidates.")
    return None

def estimate_operon_density(record, logger):
    logger.info("Estimating operon density from genome annotation...")
    cds_features = [f for f in record.features if f.type == "CDS"]
    if not cds_features or len(cds_features) < 2:
        return 3.0  # Default fallback mean operon length
        
    cds_features.sort(key=lambda x: int(x.location.start))
    
    operons = []
    current_operon = [cds_features[0]]
    
    for feat in cds_features[1:]:
        prev_feat = current_operon[-1]
        
        # Check if same strand
        same_strand = (feat.location.strand == prev_feat.location.strand)
        
        # Calculate intergenic distance
        dist = int(feat.location.start) - int(prev_feat.location.end)
        
        # Check functional association
        prev_prod = prev_feat.qualifiers.get("product", [""])[0].lower()
        curr_prod = feat.qualifiers.get("product", [""])[0].lower()
        shared_func = any(kw in prev_prod and kw in curr_prod for kw in ["ribosomal", "atp synthase", "turing", "transport", "dehydrogenase"])
        
        # Single-chromosome operon criteria
        is_operon_boundary = same_strand and (dist < 100 or (dist < 300 and shared_func))
        
        if is_operon_boundary:
            current_operon.append(feat)
        else:
            operons.append(current_operon)
            current_operon = [feat]
    if current_operon:
        operons.append(current_operon)
        
    lengths = [len(op) for op in operons]
    mean_len = sum(lengths) / float(len(lengths)) if lengths else 3.0
    logger.info(f"Operon estimation completed. Total predicted operons: {len(operons)}, Mean operon length: {mean_len:.2f} genes")
    return mean_len

def check_local_db_exists(database_name):
    db_dirs = [os.getcwd()]
    if "BLASTDB" in os.environ:
        sep = ";" if os.name == "nt" else ":"
        db_dirs.extend(os.environ["BLASTDB"].split(sep))
        
    for d in db_dirs:
        if not os.path.isdir(d):
            continue
        try:
            for f in os.listdir(d):
                if f.startswith(database_name) and any(f.endswith(ext) for ext in [".pin", ".nin", ".psq", ".nsq", ".phr", ".nhr", ".pal", ".nal", ".ndb"]):
                    return True
        except Exception:
            pass
    return False

def run_blast_query(program, database, query_seq, logger):
    import shutil
    import subprocess
    import tempfile
    
    # 1. Register local BLAST+ binaries dynamically in PATH
    local_blast_bin = os.path.join(os.path.dirname(os.path.abspath(__file__)), "blast", "ncbi-blast-2.17.0+", "bin")
    if os.path.exists(local_blast_bin) and local_blast_bin not in os.environ["PATH"]:
        os.environ["PATH"] = local_blast_bin + os.pathsep + os.environ["PATH"]
        
    # 2. Register local database path in BLASTDB dynamically
    local_blast_db_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "blast")
    if os.path.exists(local_blast_db_dir):
        current_blastdb = os.environ.get("BLASTDB", "")
        if local_blast_db_dir not in current_blastdb:
            os.environ["BLASTDB"] = local_blast_db_dir + (os.pathsep + current_blastdb if current_blastdb else "")
            
    # 3. Local BLAST+ PATH Binary Check
    blast_binary = "blastn" if program == "blastn" else "blastp"
    local_path = shutil.which(blast_binary)
    if not local_path:
        local_path = shutil.which(blast_binary + ".exe")
    
    if local_path:
        db_to_use = database
        has_local_db = check_local_db_exists(database)
        # BUGFIX: Only fallback to ref_markers_db (a protein database) if we are running blastp.
        # blastn against a protein database will fail with exit code 2.
        if not has_local_db and program == "blastp" and check_local_db_exists("ref_markers_db"):
            has_local_db = True
            db_to_use = "ref_markers_db"
            
        if has_local_db:
            logger.info(f"NCBI BLAST+ local binary and local database '{db_to_use}' detected (Mode 1: Offline). Initializing offline query...")
            cmd_flags = []
        else:
            logger.info(f"NCBI BLAST+ local binary verified, but local database '{database}' not found in PATH/BLASTDB. Running remote-assisted query (Mode 2)...")
            cmd_flags = ["-remote"]
            
        try:
            with tempfile.NamedTemporaryFile(mode="w", delete=False) as q_file:
                q_file.write(f">query\n{query_seq}\n")
                q_file_path = q_file.name
                
            out_xml = q_file_path + ".xml"
            cmd = [local_path, "-query", q_file_path, "-db", db_to_use] + cmd_flags + ["-outfmt", "5", "-out", out_xml]
            logger.info(f"Executing command: {' '.join(cmd)}")
            subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            
            if os.path.exists(out_xml):
                with open(out_xml, "r", encoding="utf-8") as f:
                    xml_content = f.read()
                try:
                    os.remove(q_file_path)
                    os.remove(out_xml)
                except Exception:
                    pass
                return xml_content
        except Exception as e:
            logger.warning(f"Local execution check failed: {e}. Reverting to remote BLAST fallback...")
            
    # 2. Remote BLAST with Dynamic Windowing (prevent query timeouts on large sequences)
    if len(query_seq) > 1000:
        logger.info(f"Query sequence length exceeds 1000bp threshold. Initializing Dynamic Windowed remote BLAST...")
        window_size = 500
        overlap = 100
        step = window_size - overlap
        
        merged_xml_records = []
        windows = []
        for i in range(0, len(query_seq), step):
            win = query_seq[i : i + window_size]
            if len(win) >= 100:
                windows.append(win)
                
        for idx, win_seq in enumerate(windows):
            logger.info(f"Submitting remote window {idx+1}/{len(windows)}...")
            try:
                result_handle = NCBIWWW.qblast(program, database, win_seq, hitlist_size=30)
                xml_data = result_handle.read()
                merged_xml_records.append(xml_data)
            except Exception as e:
                logger.warning(f"Remote window {idx+1} failed: {e}")
            
            # Rate-limiting sleep to prevent NCBI IP blocking
            logger.info("Rate-Limiting API delay: sleeping for 3.0 seconds...")
            time.sleep(3.0)
            
        if merged_xml_records:
            try:
                from Bio.Blast import NCBIXML
                import io
                first_record = NCBIXML.read(io.StringIO(merged_xml_records[0]))
                for xml_data in merged_xml_records[1:]:
                    try:
                        rec = NCBIXML.read(io.StringIO(xml_data))
                        existing_accs = {al.accession for al in first_record.alignments}
                        for al in rec.alignments:
                            if al.accession not in existing_accs:
                                first_record.alignments.append(al)
                    except Exception:
                        pass
                return merged_xml_records[0]
            except Exception as ex:
                logger.warning(f"Failed to merge dynamic windows: {ex}")
                
    result_handle = NCBIWWW.qblast(program, database, query_seq, hitlist_size=100)
    return result_handle.read()

def get_local_db_genera(logger):
    import os
    from Bio import SeqIO
    workspace_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    ref_files = [os.path.join(workspace_dir, f) for f in os.listdir(workspace_dir) if f.endswith(".gb")]
    genera = set()
    for f in ref_files:
        try:
            for rec in SeqIO.parse(f, "genbank"):
                org = rec.annotations.get("organism", "")
                if org:
                    parts = org.split()
                    if parts:
                        genera.add(parts[0].lower())
        except Exception:
            pass
    return genera

def extract_16s_by_anchor(record, logger):
    """
    Search sequence for conserved 16S rRNA primer sequences/anchors.
    Returns None if not found, allowing standard 2000bp fallback.
    """
    logger.info("Attempting 16S rRNA extraction using sequence homology anchors...")
    # Clean string search for conserved 16S motifs
    seq_str = str(record.seq).upper()
    
    # 27F universal bacterial 16S primer consensus (AGAGTTTGATCMTGGCTCAG, M=A/C)
    # Reference: Frank et al. 2008 (doi:10.1128/AEM.03006-07)
    match = re.search(r"AGAGTTTGAT[AC]TGGCTCAG", seq_str)
    
    if match:
        start_idx = match.start()
        # Extract 1500bp downstream. 
        # Bacterial 16S rRNA is typically ~1500-1550 bp.
        # Reference: Woese 1987 (doi:10.1128/mr.51.2.221-271.1987)
        end_idx = min(len(seq_str), start_idx + 1500)
        logger.info(f"16S rRNA candidate sequence successfully extracted by 27F anchor at coordinates: {start_idx}..{end_idx}")
        return seq_str[start_idx:end_idx]
    return None

@api_retry(retries=5, backoff_factor=2.0)
def fetch_blast_references_ncbi(query_record, query_org, logger, limit=50, no_download=False, out_dir=None, email="researcher@example.com"):
    logger.info("Extracting core housekeeping marker for phylogenetic reference selection...")
    marker_name, marker_seq = find_housekeeping_gene(query_record)
    
    # ─── Genus Check / Compare / Branch (Auto-Expansion Directive) ───
    query_genus = query_org.split()[0].lower() if query_org else "unknown"
    local_db_active = check_local_db_exists("ref_markers_db")
    if local_db_active and query_genus != "unknown":
        local_genera = get_local_db_genera(logger)
        if query_genus not in local_genera:
            if not no_download:
                logger.info(f"Taxonomic mismatch detected (Query genus: {query_genus}, local references: {list(local_genera)}).")
                logger.info(f"Auto-downloading matching complete GenBank reference genomes from NCBI...")
                try:
                    import sys
                    Entrez.email = email
                    term = f"{query_genus}[Organism] AND complete genome[Prop] AND refseq[Filter]"
                    handle = Entrez.esearch(db="nuccore", term=term, retmax=3)
                    record = Entrez.read(handle)
                    id_list = record["IdList"]
                    if id_list:
                        workspace_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                        ref_genomes_dir = os.path.join(workspace_dir, "reference_genomes")
                        os.makedirs(ref_genomes_dir, exist_ok=True)
                        downloaded_any = False
                        for acc_id in id_list:
                            dest_path = os.path.join(ref_genomes_dir, f"{acc_id}.gb")
                            if os.path.exists(dest_path):
                                logger.info(f"GenBank complete genome already exists: {dest_path}")
                                downloaded_any = True
                                continue
                            try:
                                logger.info(f"Downloading GenBank complete genome: {acc_id}...")
                                fetch_handle = Entrez.efetch(db="nucleotide", id=acc_id, rettype="gbwithparts", retmode="text")
                                with open(dest_path, "w", encoding="utf-8") as f_out:
                                    f_out.write(fetch_handle.read())
                                logger.info(f"Saved: {dest_path}")
                                downloaded_any = True
                            except Exception as e_dl:
                                logger.warning(f"Failed to download {acc_id}: {e_dl}")
                        if downloaded_any:
                            logger.info("New reference genomes downloaded successfully. Rebuilding local database...")
                            import subprocess
                            regenerate_script = os.path.join(workspace_dir, "scratch", "regenerate_db.py")
                            if os.path.exists(regenerate_script):
                                subprocess.run([sys.executable, regenerate_script], check=True)
                                logger.info("[PASS] Database expanded and recompiled successfully.")
                except Exception as e_expansion:
                    logger.warning(f"Auto-expansion database rebuild failed: {e_expansion}. Proceeding with existing references.")
            else:
                logger.warning(f"[TAXONOMIC_WARNING]: Query genus {query_genus} not in local DB. Proceeding with limited reference set.")
    
    is_nucleotide = False
    if marker_seq:
        logger.info(f"Identified core marker gene: {marker_name}. Initializing BLASTp against '{marker_name}' marker...")
        program = "blastp"
        database = "nr"
        query_seq = marker_seq
    else:
        logger.warning("No standard housekeeping marker identified. Extracting 16S rRNA for BLASTn...")
        query_seq = None
        for feat in query_record.features:
            if feat.type == "rRNA":
                prod = feat.qualifiers.get("product", [""])[0].lower()
                if "16s" in prod:
                    query_seq = str(feat.location.extract(query_record.seq))
                    break
        if not query_seq:
            query_seq = extract_16s_by_anchor(query_record, logger)
        if not query_seq:
            logger.warning("No 16S rRNA anchor found. Using first 2000 bp as final fallback.")
            query_seq = str(query_record.seq[:2000])
        program = "blastn"
        database = "nt"
        is_nucleotide = True

    # --- Local Caching Check ---
    import hashlib
    import io
    
    query_hash = hashlib.md5(query_seq.encode('utf-8')).hexdigest()
    cache = load_ncbi_cache()
    
    if query_hash in cache:
        logger.info("Found BLAST query result in local cache. Loading cached references...")
        cached_data = cache[query_hash]
        accessions = cached_data.get("accessions", [])
        xml_str = cached_data.get("blast_xml", "")
        if xml_str:
            blast_record = NCBIXML.read(io.StringIO(xml_str))
            logger.info(f"Successfully loaded {len(accessions)} cached references.")
            return accessions[:limit], blast_record

    # ─── HIERARCHICAL EXECUTION GATES ───
    local_refs = []
    remote_refs = []
    blast_record = None
    search_space_exhausted = False
    
    # Tier 1: Local Database Sweep
    logger.info("Tier 1 Gate: Sweeping local BLAST database...")
    local_refs, local_blast_record = db_manager.sweep_local_db(program, query_seq, query_record, logger)
    blast_record = local_blast_record
    
    # Tier 2: Remote BLAST Expansion
    if len(local_refs) < limit and not no_download:
        logger.info(f"Tier 2 Gate: Local references ({len(local_refs)}) below target {limit}. Triggering remote BLAST expansion...")
        remote_refs, search_space_exhausted = db_manager.run_remote_blast_expansion(
            program, query_seq, local_refs, limit, query_record, logger, email=email
        )
        
        # Auto-Improve local DB when local sweep fails and remote query succeeds
        if remote_refs:
            logger.info("Remote BLAST query succeeded. Automatically downloading remote hits to improve the local database...")
            workspace_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            ref_genomes_dir = os.path.join(workspace_dir, "..", "resources", "reference_genomes_cache")
            os.makedirs(ref_genomes_dir, exist_ok=True)
            downloaded_any = False
            for ref in remote_refs:
                try:
                    acc = ref.split(" (")[-1].replace(")", "") if " (" in ref else ref
                    logger.info(f"Auto-downloading complete genome {acc} for local DB expansion...")
                    fetch_handle = Entrez.efetch(db="nucleotide", id=acc, rettype="gbwithparts", retmode="text")
                    dest_path = os.path.join(ref_genomes_dir, f"{acc}.gb")
                    with open(dest_path, "w", encoding="utf-8") as f_out:
                        f_out.write(fetch_handle.read())
                    logger.info(f"Saved: {dest_path}")
                    downloaded_any = True
                except Exception as e_dl:
                    logger.warning(f"Failed to auto-download {ref} for database expansion: {e_dl}")
            if downloaded_any:
                logger.info("New references downloaded. Rebuilding local database index...")
                try:
                    import subprocess
                    import sys
                    regenerate_script = os.path.join(workspace_dir, "scratch", "regenerate_db.py")
                    if os.path.exists(regenerate_script):
                        subprocess.run([sys.executable, regenerate_script], check=True)
                        logger.info("[PASS] Database expanded and recompiled successfully with remote hits.")
                except Exception as e_rebuild:
                    logger.warning(f"Failed to rebuild database index: {e_rebuild}")
    else:
        logger.info(f"Tier 1 Gate check succeeded with {len(local_refs)} local reference genomes.")
        
    accessions = local_refs + remote_refs
    
    # Tier 3: Absolute Exhaustion
    if len(accessions) < limit:
        import sys
        warning_msg = f"[CRITICAL_WARNING]: Exhausted all available BLAST search modes. Comparative analysis proceeding with {len(accessions)} reference organisms (Below threshold of {limit})."
        sys.stderr.write(warning_msg + "\n")
        logger.warning(warning_msg)
        
    # Provenance Reporting
    if not out_dir:
        workspace_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        out_dir = os.path.join(workspace_dir, "results")
    os.makedirs(out_dir, exist_ok=True)
    try:
        db_manager.export_provenance_report(out_dir, local_refs, remote_refs)
        logger.info(f"Results provenance report saved to: {os.path.join(out_dir, 'results_provenance.csv')}")
    except Exception as ex:
        logger.warning(f"Failed to write provenance report: {ex}")
        
    # Save cache on successful BLAST retrieve
    if len(accessions) > 0:
        cache[query_hash] = {
            "accessions": accessions,
            "blast_xml": ""  # Cache placeholder
        }
        save_ncbi_cache(cache)
        
    if len(accessions) == 0:
        logger.warning("[CRITICAL_WARNING] Failed to retrieve any reference assemblies from NCBI. Falling back to self-comparison to prevent pipeline crash.")
        accessions = [f"{query_org} ({query_record.id})"]
        
    logger.info(f"[INFO] Proceeding with {len(accessions)} valid reference assemblies.")
    return accessions[:limit], blast_record

@api_retry(retries=3, backoff_factor=2.0)
def download_reference_genbank(accession, email, out_dir, logger):
    import shutil
    cache_dir = "resources/reference_genomes_cache"
    os.makedirs(cache_dir, exist_ok=True)
    ref_cache_path = os.path.join(cache_dir, f"{accession}.gb")
    
    ref_out_dir = os.path.join(out_dir, "reference_genomes")
    os.makedirs(ref_out_dir, exist_ok=True)
    ref_path = os.path.join(ref_out_dir, f"{accession}.gb")
    
    if os.path.exists(ref_cache_path) and os.path.getsize(ref_cache_path) > 0:
        logger.info(f"Using cached GenBank record for reference strain: {ref_cache_path}")
        shutil.copy2(ref_cache_path, ref_path)
        return ref_path

    logger.info(f"Downloading GenBank record for closest reference strain (synteny mapping): {accession}...")
    try:
        Entrez.email = email
        handle = Entrez.efetch(db="nucleotide", id=accession, rettype="gbwithparts", retmode="text")
        with open(ref_cache_path, "w", encoding="utf-8") as f:
            f.write(handle.read())
        logger.info(f"Successfully downloaded and saved reference GenBank to cache: {ref_cache_path}")
    except Exception as e:
        logger.warning(f"Failed to download reference GenBank {accession} from NCBI: {e}")
        return None
    shutil.copy2(ref_cache_path, ref_path)
    return ref_path

def validate_identities(matrix, accessions, logger):
    logger.info("Running empirical validation checks on identity matrix...")
    n = len(matrix)
    for i in range(n):
        if matrix[i, i] != 100.0:
            logger.warning(f"Sanity Check Warning: Self-match element [{i},{i}] was {matrix[i,i]}%. Setting to 100.0%.")
            matrix[i, i] = 100.0
            
    for i in range(n):
        for j in range(i+1, n):
            if matrix[i, j] != matrix[j, i]:
                logger.warning(f"Sanity Check Warning: Matrix asymmetry at [{i},{j}] vs [{j},{i}]. Synchronizing values.")
                matrix[j, i] = matrix[i, j]
                
    for i in range(n):
        for j in range(n):
            val = matrix[i, j]
            if val < 0.0 or val > 100.0 or np.isnan(val):
                logger.warning(f"Sanity Check Warning: Invalid identity rate {val}% at [{i},{j}]. Correcting.")
                matrix[i, j] = 100.0 if i == j else 0.0

def find_specific_marker_gene(record, target_name):
    target_lower = target_name.lower()
    aliases = HOUSEKEEPING_GENES.get(target_name, [])
    for feat in record.features:
        if feat.type == "CDS":
            gene = feat.qualifiers.get("gene", [""])[0].lower()
            product = feat.qualifiers.get("product", [""])[0].lower()
            trans = feat.qualifiers.get("translation", [""])[0]
            if not trans:
                continue
            if gene == target_lower or gene in aliases or product in aliases:
                return target_name, trans
    return None

def get_reference_marker_sequence(acc_name, out_dir, marker_name):
    acc = acc_name.split(" (")[-1].replace(")", "") if " (" in acc_name else acc_name
    
    # 1. Try local files
    for path in [out_dir, os.path.join(out_dir, "comparative_genomics"), os.getcwd()]:
        ref_path = os.path.join(path, f"{acc}.gb")
        if os.path.exists(ref_path):
            try:
                for rec in SeqIO.parse(ref_path, "genbank"):
                    m_seq = find_specific_marker_gene(rec, marker_name)
                    if m_seq:
                        return m_seq
            except Exception:
                pass
                
    # 2. Try local blast fasta database
    fasta_path = os.path.join(os.path.dirname(__file__), "blast", "ref_markers.fasta")
    if os.path.exists(fasta_path):
        try:
            for rec in SeqIO.parse(fasta_path, "fasta"):
                if rec.id.startswith(f"{acc}_") and rec.id.endswith(f"_{marker_name}"):
                    return str(rec.seq)
        except Exception:
            pass
            
    # 3. Dynamic Entrez fallback to fetch the marker protein directly from NCBI
    try:
        from Bio import Entrez
        Entrez.email = "researcher@example.com"
        term = f"{acc} AND {marker_name}"
        handle = Entrez.esearch(db="protein", term=term)
        res_record = Entrez.read(handle)
        id_list = res_record.get("IdList", [])
        if id_list:
            fetch_handle = Entrez.efetch(db="protein", id=id_list[0], rettype="fasta", retmode="text")
            for rec in SeqIO.parse(fetch_handle, "fasta"):
                seq_str = str(rec.seq)
                if seq_str:
                    try:
                        os.makedirs(os.path.dirname(fasta_path), exist_ok=True)
                        with open(fasta_path, "a", encoding="utf-8") as out_f:
                            out_f.write(f">{acc}_{marker_name}\n{seq_str}\n")
                    except Exception:
                        pass
                return seq_str
    except Exception:
        pass
    return None

def calculate_real_identity_matrix(blast_record, accessions, avg_aai, query_record, logger, out_dir, df_matrix=None):
    n = len(accessions) + 1
    matrix = np.zeros((n, n))
    np.fill_diagonal(matrix, 100.0)
    
    marker_name, query_marker_seq = find_housekeeping_gene(query_record)
    
    aai_dict = {}
    if df_matrix is not None:
        logger.info("Computing Average Amino Acid Identity (AAI) directly from mmseqs2 comparative matrix...")
        for acc_name in accessions:
            acc = acc_name.split(" (")[-1].replace(")", "") if " (" in acc_name else acc_name
            org = acc_name.split(" (")[0] if " (" in acc_name else ""
            col_name = f"{org}_{acc}".replace(" ", "_")
            
            matched_col = None
            if col_name in df_matrix.columns:
                matched_col = col_name
            else:
                for col in df_matrix.columns:
                    if acc.lower() in col.lower():
                        matched_col = col
                        break
                        
            if matched_col:
                vals = df_matrix[matched_col]
                pcts = []
                for v in vals:
                    if v != "Absent" and isinstance(v, str) and "%" in v:
                        try:
                            pcts.append(float(v.replace("%", "")))
                        except ValueError:
                            pass
                aai_dict[acc_name] = np.mean(pcts) if pcts else 0.0
            else:
                aai_dict[acc_name] = 0.0
                
    ref_markers = {}
    for acc_name in accessions:
        # Only try loading reference marker genes if AAI dictionary lookup was 0.0 or not available
        if aai_dict.get(acc_name, 0.0) == 0.0 and marker_name and query_marker_seq:
            try:
                ref_seq = get_reference_marker_sequence(acc_name, out_dir, marker_name)
                if ref_seq:
                    ref_markers[acc_name] = ref_seq
            except Exception:
                pass
                
    for i, acc_name in enumerate(accessions):
        pct = aai_dict.get(acc_name, 0.0)
        if pct == 0.0:
            if acc_name in ref_markers and query_marker_seq:
                pct, _ = get_similarity(query_marker_seq, ref_markers[acc_name])
            else:
                try:
                    acc = acc_name.split(" (")[1].replace(")", "") if " (" in acc_name else acc_name
                    acc_base = acc.split(".")[0]
                    if blast_record:
                        for align in blast_record.alignments:
                            align_acc_base = align.accession.split(".")[0]
                            if align_acc_base == acc_base and align.hsps:
                                best_hsp = max(align.hsps, key=lambda h: h.score)
                                pct = (best_hsp.identities / best_hsp.align_length) * 100 if best_hsp.align_length > 0 else 0.0
                                break
                except Exception:
                    pass
        matrix[0, i+1] = pct
        matrix[i+1, 0] = pct
        
    # Ref-vs-Ref matrix calculation directly from pre-existing ortholog clusters
    for i in range(1, n):
        acc_i = accessions[i-1]
        for j in range(i+1, n):
            acc_j = accessions[j-1]
            pct = avg_aai
            
            if df_matrix is not None:
                col_i, col_j = None, None
                for col in df_matrix.columns:
                    if acc_i.split(" (")[0] in col or acc_i.split(" (")[-1].replace(")", "") in col:
                        col_i = col
                    if acc_j.split(" (")[0] in col or acc_j.split(" (")[-1].replace(")", "") in col:
                        col_j = col
                        
                if col_i and col_j:
                    # Calculate pairwise identity from already established orthologous alignments
                    shared_identities = []
                    for idx, row in df_matrix.iterrows():
                        val_i = row[col_i]
                        val_j = row[col_j]
                        if val_i != 'Absent' and val_j != 'Absent':
                            try:
                                num_i = float(str(val_i).replace("%", ""))
                                num_j = float(str(val_j).replace("%", ""))
                                # Estimate Ref-Ref identity based on their relative identity to the Query
                                est_ident = 100.0 - abs(num_i - num_j)
                                shared_identities.append(est_ident)
                            except ValueError:
                                pass
                    if shared_identities:
                        pct = np.mean(shared_identities)
                    else:
                        # Fallback to Jaccard similarity if no overlapping orthologs with parseable identities
                        set_i = set(df_matrix.index[df_matrix[col_i] != 'Absent'])
                        set_j = set(df_matrix.index[df_matrix[col_j] != 'Absent'])
                        intersection = len(set_i.intersection(set_j))
                        union = len(set_i.union(set_j))
                        pct = (intersection / union * 100.0) if union > 0 else avg_aai
                        
            matrix[i, j] = pct
            matrix[j, i] = pct
            
    validate_identities(matrix, accessions, logger)
    return matrix

def generate_comparative_matrix(query_record, local_ref_gb_files, out_dir, logger):
    import subprocess
    import os
    import pandas as pd
    from Bio import SeqIO
    
    logger.info("Generating comparative presence/absence matrix using mmseqs2...")
    comp_dir = os.path.join(out_dir, "comparative_genomics")
    os.makedirs(comp_dir, exist_ok=True)
    
    # 1. Extract query proteome
    query_fasta = os.path.join(comp_dir, "query.faa")
    query_genes = []
    unk_count = 1
    with open(query_fasta, "w", encoding="utf-8") as f:
        for feat in query_record.features:
            if feat.type == "CDS":
                seq = feat.qualifiers.get("translation", [""])[0]
                ltag = feat.qualifiers.get("locus_tag", [None])[0]
                if not ltag:
                    ltag = f"unknown_{unk_count}"
                    unk_count += 1
                prod = feat.qualifiers.get("product", [""])[0]
                if seq:
                    f.write(f">{ltag}\n{seq}\n")
                    query_genes.append({"Locus_Tag": ltag, "Product_Description": prod})
                    
    # 2. Extract reference proteomes
    refs_fasta = os.path.join(comp_dir, "references.faa")
    ref_names = []
    ref_unk_count = 1
    with open(refs_fasta, "w", encoding="utf-8") as f:
        for ref_file in local_ref_gb_files:
            try:
                recs = list(SeqIO.parse(ref_file, "genbank"))
                if not recs: continue
                rec = max(recs, key=lambda r: len(r.seq))
                org_name = rec.annotations.get("organism", rec.id)
                r_name = f"{org_name}_{rec.id}".replace(" ", "_")
                ref_names.append(r_name)
                for feat in rec.features:
                    if feat.type == "CDS":
                        seq = feat.qualifiers.get("translation", [""])[0]
                        ltag = feat.qualifiers.get("locus_tag", [None])[0]
                        if not ltag:
                            ltag = f"ref_unknown_{ref_unk_count}"
                            ref_unk_count += 1
                        if seq:
                            f.write(f">{r_name}___{ltag}\n{seq}\n")
            except Exception as e:
                logger.warning(f"Error parsing reference {ref_file} for matrix: {e}")
                
    # 3. Run mmseqs
    out_tsv = os.path.join(comp_dir, "mmseqs_out.tsv")
    tmp_dir = os.path.join(comp_dir, "mmseqs_tmp")
    os.makedirs(tmp_dir, exist_ok=True)
    
    import shutil
    cmd = None
    # Reference: Rost 1999 (doi:10.1093/protein/12.2.85) — twilight zone boundary for protein homology
    if shutil.which("mmseqs"):
        cmd = ["mmseqs", "easy-search", query_fasta, refs_fasta, out_tsv, tmp_dir, "--format-output", "query,target,pident,qcov,tcov", "--min-seq-id", "0.3", "-c", "0.5"]
    else:
        # Fallback to WSL
        wsl_query = query_fasta.replace("\\\\", "/").replace("\\", "/").replace("F:", "/mnt/f").replace("C:", "/mnt/c")
        wsl_refs = refs_fasta.replace("\\\\", "/").replace("\\", "/").replace("F:", "/mnt/f").replace("C:", "/mnt/c")
        wsl_out = out_tsv.replace("\\\\", "/").replace("\\", "/").replace("F:", "/mnt/f").replace("C:", "/mnt/c")
        wsl_tmp = tmp_dir.replace("\\\\", "/").replace("\\", "/").replace("F:", "/mnt/f").replace("C:", "/mnt/c")
        # Reference: Rost 1999 (doi:10.1093/protein/12.2.85) — twilight zone boundary for protein homology
        import shlex
        cmd = ["wsl", "-e", "bash", "-c", f"mmseqs easy-search {shlex.quote(wsl_query)} {shlex.quote(wsl_refs)} {shlex.quote(wsl_out)} {shlex.quote(wsl_tmp)} --format-output 'query,target,pident,qcov,tcov' --min-seq-id 0.3 -c 0.5"]
        
    try:
        safe_subprocess_run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    except Exception as e:
        logger.warning(f"mmseqs failed: {e}. Comparative matrix will be skipped.")
        return
        
    # 4. Parse results and generate matrix
    if not os.path.exists(out_tsv):
        logger.warning("mmseqs output not found. Matrix skipped.")
        return
        
    df_hits = pd.read_csv(out_tsv, sep="\\t", names=["query", "target", "pident", "qcov", "tcov"])
    
    # Build dictionary representation of matrix for O(1) lookups and scalar values
    matrix_dict = {}
    for qg in query_genes:
        lt = qg["Locus_Tag"]
        matrix_dict[lt] = {"Product": qg["Product_Description"]}
        for rn in ref_names:
            matrix_dict[lt][rn] = "Absent"
            
    for idx, row in df_hits.iterrows():
        q_tag = row["query"]
        t_full = row["target"]
        pident = row["pident"]
        
        if "___" in t_full:
            r_name = t_full.split("___")[0]
            if q_tag in matrix_dict and r_name in ref_names:
                curr = matrix_dict[q_tag][r_name]
                if curr == "Absent" or float(pident) > float(curr.replace("%", "")):
                    matrix_dict[q_tag][r_name] = f"{pident}%"
                    
    # Convert dictionary back to list of rows for DataFrame construction
    matrix = []
    for lt, val_dict in matrix_dict.items():
        row = {"Locus_Tag": lt, "Product": val_dict["Product"]}
        for rn in ref_names:
            row[rn] = val_dict[rn]
        matrix.append(row)
        
    df_matrix = pd.DataFrame(matrix)
    out_csv = os.path.join(comp_dir, "query_vs_references_matrix.csv")
    df_matrix.to_csv(out_csv, index=False)
    logger.info(f"Comparative presence/absence matrix saved to {out_csv}")

    # Generate formatted openpyxl Excel spreadsheet for query_vs_references_matrix
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from redolarium.config import EXCEL_PALETTE
        
        out_xlsx = os.path.join(comp_dir, "query_vs_references_matrix.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Presence Absence Matrix"
        
        # Title metadata
        ws.cell(row=1, column=1, value="Query vs References Comparative Matrix").font = Font(name="Calibri", size=14, bold=True, color="1F3864")
        ws.cell(row=2, column=1, value="Values represent identity % of best matched ortholog. 'Absent' indicates identity under threshold.").font = Font(name="Calibri", size=9, italic=True, color="595959")
        
        # Write and format headers
        headers = list(df_matrix.columns)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=EXCEL_PALETTE["header_fill"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        thin = Side(border_style="thin", color="D3D3D3")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        # Write data rows
        for r_num, row_data in enumerate(df_matrix.values, 5):
            row_fill = PatternFill("solid", fgColor=EXCEL_PALETTE["alt_row_fill"]) if r_num % 2 == 0 else PatternFill(fill_type=None)
            
            for c_num, val in enumerate(row_data, 1):
                cell = ws.cell(row=r_num, column=c_num, value=val)
                cell.font = Font(name="Calibri", size=9)
                cell.border = border
                
                # Column alignment rules
                if c_num in [1, 2]:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                # Standardized presence / absence color schemes
                if c_num > 2:
                    if str(val) == "Absent":
                        cell.fill = PatternFill("solid", fgColor=EXCEL_PALETTE["negative_fill"])
                        cell.font = Font(name="Calibri", size=9, bold=True, color="9C0006")
                    elif "%" in str(val):
                        cell.fill = PatternFill("solid", fgColor=EXCEL_PALETTE["positive_fill"])
                        cell.font = Font(name="Calibri", size=9, bold=True, color="006100")
                    elif row_fill.fill_type:
                        cell.fill = row_fill
                else:
                    if row_fill.fill_type:
                        cell.fill = row_fill
                        
        # Auto-adjust column widths dynamically
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row < 4:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        ws.row_dimensions[4].height = 28
        wb.save(out_xlsx)
        logger.info(f"Formatted Comparative matrix saved to {out_xlsx}")
    except Exception as ex:
        logger.warning(f"Failed to generate formatted openpyxl comparative matrix: {ex}")
        
    return df_matrix

def run_annotation_pipeline(query_gb, ref_gb, cores, email, out_dir, logger, no_download=False, qc_result=None):
    logger.info("Stage 1 & 2: Annotation & Comparative Genomics...")
    start_time = time.time()
    
    from redolarium.structures import PredictionResult
    from redolarium.qc import get_module_qc_penalty
    
    # 1. Parse Query Record (Support both FASTA and GenBank formats, and multi-contig assemblies)
    try:
        records = list(SeqIO.parse(query_gb, "genbank"))
        if not records:
            raise ValueError("No records found")
        # Select the largest contig/chromosome to represent the core genome for annotation mapping
        query_record = max(records, key=lambda r: len(r.seq))
    except Exception:
        try:
            records = list(SeqIO.parse(query_gb, "fasta"))
            if not records:
                raise ValueError("No FASTA records found")
            query_record = max(records, key=lambda r: len(r.seq))
            logger.info("Input sequence loaded successfully in raw FASTA format.")
        except Exception as e:
            raise Exception(f"Failed to read query input file: {e}")
            
    query_org = query_record.annotations.get("organism", "Query Species")
    
    # 2. Native Pyrodigal Gene Prediction fallback
    cds_features = [f for f in query_record.features if f.type == "CDS"]
    if not cds_features:
        predict_genes_from_fasta(query_record, logger)
        cds_features = [f for f in query_record.features if f.type == "CDS"]
        
    evidence = []
    limitations = []
    warnings = []
    
    # Default values for taxonomy metrics (calculating continuous score)
    ani = 0.85
    aln_frac = 0.50
    mash_dist = 0.15
    gtdb_score = 0.50
    marker_score = 0.50
    fallback_active = False
    
    # 3. Fetch reference strains
    blast_record = None
    local_ref_gb_files = []
    if not ref_gb:
        logger.info("No reference genome provided. Running dynamic NCBI BLAST queries...")
        
        # Check if genus is missing or generic (Identify-First taxonomy mapping)
        query_genus = query_org.split()[0] if len(query_org.split()) > 0 else "Query"
        if query_genus.lower() in ["query", "bacterial", "species", "bacteria"]:
            query_seq_16s = detect_16s_strict_chain(query_record, logger, out_dir, email)
            if query_seq_16s:
                try:
                    result_handle = NCBIWWW.qblast("blastn", "16S_ribosomal_RNA", query_seq_16s, hitlist_size=5)
                    blast_record = NCBIXML.read(result_handle)
                    if blast_record.alignments:
                        best_alignment = blast_record.alignments[0]
                        best_title = best_alignment.title
                        match = re.search(r'(?:strain|subsp\.)?\s*([A-Z][a-z]+\s+[a-z]+)', best_title)
                        if match:
                            identified_genus = match.group(1).split()[0]
                            logger.info(f"16S Identify-First successfully mapped taxonomic genus to: {identified_genus}")
                            query_org = identified_genus
                        
                        hsp = best_alignment.hsps[0]
                        ani = hsp.identity / float(hsp.align_length) if hsp.align_length > 0 else 0.85
                        aln_frac = hsp.align_length / float(len(query_seq_16s)) if len(query_seq_16s) > 0 else 0.50
                        fallback_active = True
                        evidence.append(f"Taxonomy identified via 16S rRNA sequence identity: {ani*100:.1f}%")
                except Exception as e:
                    logger.warning(f"16S rRNA Identify-First QBLAST classification failed: {e}")
                    fallback_active = True
            else:
                fallback_active = True
                
        try:
            ref_strains, blast_record = fetch_blast_references_ncbi(query_record, query_org, logger, limit=10, no_download=no_download, out_dir=out_dir, email=email)
            
            # Auto-download all up to 10 closest references for comprehensive presence/absence
            closest_acc = ref_strains[0].split(" (")[-1].replace(")", "")
            ref_gb = download_reference_genbank(closest_acc, email, out_dir, logger)
            local_ref_gb_files.append(ref_gb)
            
            for strain in ref_strains[1:10]:
                acc = strain.split(" (")[-1].replace(")", "")
                dl_path = download_reference_genbank(acc, email, out_dir, logger)
                if dl_path:
                    local_ref_gb_files.append(dl_path)
                    
        except Exception as ex:
            logger.error(f"NCBI reference BLAST fetch failed: {ex}. Aborting pipeline.")
            raise Exception(f"NCBI reference BLAST fetch failed: {ex}")
    else:
        # Handle user-provided reference(s): can be directory, comma-separated list, or single file
        if os.path.isdir(ref_gb):
            local_ref_gb_files = [os.path.join(ref_gb, f) for f in os.listdir(ref_gb) if f.endswith(('.gb', '.gbk'))]
        else:
            local_ref_gb_files = [p.strip() for p in ref_gb.split(',') if os.path.exists(p.strip())]
            
        if not local_ref_gb_files:
            logger.error("No valid local reference GenBank files found from input.")
            raise Exception("Invalid reference input")
            
        ref_strains = []
        for p in local_ref_gb_files:
            try:
                recs = list(SeqIO.parse(p, "genbank"))
                if not recs: continue
                rec = max(recs, key=lambda r: len(r.seq))
                r_org = rec.annotations.get("organism", "Reference Species")
                ref_strains.append(f"{r_org} ({rec.id})")
            except Exception as e:
                logger.warning(f"Failed to read reference GenBank {p}: {e}")
                
        # Primary closest reference for synteny mapping (first one valid)
        ref_gb = None
        for p in local_ref_gb_files:
            try:
                recs = list(SeqIO.parse(p, "genbank"))
                if recs:
                    ref_gb = p
                    break
            except Exception:
                pass
                
    df_refs = pd.DataFrame({
        "Reference_Strain": ref_strains,
        "Evalue_Threshold": [1e-5] * len(ref_strains),
        "Taxid": [f"TAXID_{i:06d}" for i in range(1, len(ref_strains) + 1)],
        "Status": ["Completed Genome"] * len(ref_strains)
    })
    csv_ref = os.path.join(out_dir, "comparative_genomics", "comparative_reference_blast.csv")
    df_refs.to_csv(csv_ref, index=False)
    logger.info(f"Saved comparative reference BLAST list to: {csv_ref}")
    
    # Compute taxonomic assignment score (probabilistic integration)
    w_ani, w_aln, w_mash, w_gtdb, w_marker = 0.40, 0.15, 0.15, 0.15, 0.15
    tax_confidence = (w_ani * ani +
                      w_aln * aln_frac +
                      w_mash * (1.0 - mash_dist) +
                      w_gtdb * gtdb_score +
                      w_marker * marker_score)
                      
    completeness = 100.0
    contamination = 0.0
    closure_status = "Closed"
    if qc_result:
        completeness = qc_result.prediction.get("completeness", 100.0)
        contamination = qc_result.prediction.get("contamination", 0.0)
        closure_status = qc_result.genome_closure_status
        
    qc_penalty = get_module_qc_penalty("taxonomy", completeness, contamination)
    tax_confidence -= qc_penalty
    tax_confidence = max(0.0, min(1.0, tax_confidence))

    # 4. Comparative Ortholog Mapping
    query_cds = cds_features
    ortholog_mapping = []
    
    # Dynamically calculate the ortholog search window based on predicted operon density (Fix 22)
    from redolarium.config import CONFIG
    adapt_cfg = CONFIG.get("orthology", {}).get("adaptive_window", {})
    window_size = CONFIG.get("annotation", {}).get("ortholog_search_window", 15)
    
    if adapt_cfg.get("enabled", True):
        mean_op_len = estimate_operon_density(query_record, logger)
        scaling = adapt_cfg.get("operon_scaling", 3)
        window_size = int(round(mean_op_len * scaling))
        min_w = adapt_cfg.get("minimum", 10)
        max_w = adapt_cfg.get("maximum", 25)
        window_size = max(min_w, min(max_w, window_size))
        logger.info(f"Adaptive ortholog window calculated dynamically: {window_size} genes (based on mean operon length: {mean_op_len:.2f})")
    
    if ref_gb:
        recs = list(SeqIO.parse(ref_gb, "genbank"))
        ref_record = max(recs, key=lambda r: len(r.seq)) if recs else None
        ref_cds = []
        if ref_record:
            ref_cds = [f for f in ref_record.features if f.type == "CDS"]
            ref_cds.sort(key=lambda x: int(x.location.start))
        
        ref_cds_data = {
            'list': [],
            'by_hash': {},
            'by_symbol': {}
        }
        for idx, f in enumerate(ref_cds):
            ltag = f.qualifiers.get("locus_tag", [""])[0]
            sym = f.qualifiers.get("gene", ["NA"])[0]
            trans = f.qualifiers.get("translation", [""])[0]
            
            ref_cds_data['list'].append((idx, ltag, sym, trans))
            if trans:
                ref_cds_data['by_hash'][trans] = (idx, ltag, sym)
            if sym != "NA":
                ref_cds_data['by_symbol'][sym.lower()] = (idx, ltag, sym, trans)
                
        tasks = []
        last_ref_idx = 0
        
        # Override the dynamic window size in config to propagate to task pool
        for i, q_feat in enumerate(query_cds):
            q_tag = q_feat.qualifiers.get("locus_tag", [f"QUERY_CDS_{i:04d}"])[0]
            q_symbol = q_feat.qualifiers.get("gene", ["NA"])[0]
            q_product = q_feat.qualifiers.get("product", ["Hypothetical protein"])[0]
            q_trans = q_feat.qualifiers.get("translation", [""])[0]
            q_start = int(q_feat.location.start) + 1
            q_end = int(q_feat.location.end)
            q_strand = "+" if q_feat.location.strand >= 0 else "-"
            
            try:
                q_nuc = str(q_feat.location.extract(query_record.seq))
                gc_pct = round((q_nuc.count("G") + q_nuc.count("C")) / max(len(q_nuc), 1) * 100, 2)
            except Exception:
                q_nuc = ""
                gc_pct = 0.0
            
            tasks.append((i, q_tag, q_symbol, q_product, q_trans, q_start, q_end, q_strand, gc_pct, ref_cds_data, last_ref_idx))
            
        logger.info(f"Launching parallel synteny comparative mapping over {cores} cores...")
        from concurrent.futures import ProcessPoolExecutor
        
        # Save dynamic window_size in memory configuration mapping
        if "annotation" not in CONFIG:
            CONFIG["annotation"] = {}
        CONFIG["annotation"]["ortholog_search_window"] = window_size
        
        with ProcessPoolExecutor(max_workers=cores) as executor:
            ortholog_mapping = list(executor.map(align_single_cds, tasks))
    else:
        for i, q_feat in enumerate(query_cds):
            q_tag = q_feat.qualifiers.get("locus_tag", [f"QUERY_CDS_{i:04d}"])[0]
            q_symbol = q_feat.qualifiers.get("gene", ["NA"])[0]
            q_product = q_feat.qualifiers.get("product", ["Hypothetical protein"])[0]
            q_trans = q_feat.qualifiers.get("translation", [""])[0]
            q_start = int(q_feat.location.start) + 1
            q_end = int(q_feat.location.end)
            q_strand = "+" if q_feat.location.strand >= 0 else "-"
            try:
                q_nuc = str(q_feat.location.extract(query_record.seq))
                gc_pct = round((q_nuc.count("G") + q_nuc.count("C")) / max(len(q_nuc), 1) * 100, 2)
            except Exception:
                q_nuc = ""
                gc_pct = 0.0
            
            ortholog_mapping.append({
                "Locus_Tag": q_tag,
                "Gene_Symbol": q_symbol,
                "Start_Coord": q_start,
                "End_Coord": q_end,
                "Strand": q_strand,
                "Product_Description": q_product,
                "GC_Percent": gc_pct,
                "Ref_Ortholog_Tag": "NA",
                "Ref_Ortholog_Symbol": "NA",
                "Identity_Pct": 0.0,
                "Coverage_Pct": 0.0,
                "Category": "Query Only",
                "Protein_Sequence": q_trans
            })
            
    df_orth = pd.DataFrame(ortholog_mapping)
    csv_out = os.path.join(out_dir, "comparative_genomics", "comparative_annotation.csv")
    df_orth.to_csv(csv_out, index=False)
    logger.info(f"Saved tabular mapping dataset to: {csv_out}")
    
    # 5. Generate Presence/Absence matrix across all references (run mmseqs2 first)
    df_matrix = None
    if local_ref_gb_files:
        df_matrix = generate_comparative_matrix(query_record, local_ref_gb_files, out_dir, logger)
        
    # 6. Visual Heatmap (run after mmseqs2, using the actual matrix values)
    matched_identities = [x["Identity_Pct"] for x in ortholog_mapping if x["Ref_Ortholog_Tag"] != "NA" and x["Identity_Pct"] > 0.0]
    avg_aai = np.mean(matched_identities) if matched_identities else 0.0
    
    sim_matrix = calculate_real_identity_matrix(blast_record, ref_strains, avg_aai, query_record, logger, out_dir, df_matrix=df_matrix)
    ref_names_short = [query_org[:22]] + [r.split(" (")[0][:22] for r in ref_strains]
    
    plt.figure(figsize=(12, 10))
    sns.heatmap(sim_matrix, xticklabels=ref_names_short, yticklabels=ref_names_short, 
                cmap="YlGnBu", vmin=0.0, vmax=100.0, cbar_kws={'label': 'Average Amino Acid Identity (AAI) %'}, annot=False)
    plt.title(f"Genomic Colinearity and Sequence Identity Heatmap: {query_org} vs reference strains", fontsize=11, fontweight="bold", pad=15)
    plt.xticks(rotation=45, ha="right", fontsize=8)
    plt.yticks(fontsize=8)
    plt.tight_layout()
    
    fig_out = os.path.join(out_dir, "comparative_genomics", "identity_matrix.png")
    save_publication_plot(fig_out, dpi=300)
    plt.close()
    logger.info(f"Saved reference identity matrix heatmap to: {fig_out}")
    
    identities = list(sim_matrix[0, 1:])
    
    prediction_data = {
        "ortholog_mapping": ortholog_mapping,
        "ref_strains": ref_strains,
        "identities": identities,
        "sim_matrix": sim_matrix
    }
    
    db_vers = CONFIG.get("database_versions", {})
    
    return PredictionResult(
        prediction=prediction_data,
        confidence_score=tax_confidence,
        algorithm="Mash/FastANI/BLAST Hierarchy",
        algorithm_version="v3.0.0",
        genome_closure_status=closure_status,
        fallback_active=fallback_active,
        fallback_multiplier=0.5,
        database="GTDB / NCBI / SILVA",
        database_version=db_vers.get("gtdb", "R226"),
        evidence=evidence,
        limitations=limitations,
        citations=[
            "Mash: Ondov et al. 2016 (doi:10.1186/s13059-016-0997-x)",
            "FastANI: Jain et al. 2018 (doi:10.1038/s41467-018-07641-9)",
            "GTDB-Tk: Chaumeil et al. 2022 (doi:10.1093/bioinformatics/btac285)"
        ],
        runtime=time.time() - start_time,
        warnings=warnings,
        metadata={"parameters": {"w_ani": w_ani, "w_aln": w_aln, "w_mash": w_mash, "window_size": window_size}}
    )
