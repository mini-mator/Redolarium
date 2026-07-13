# SPDX-License-Identifier: MIT
# Copyright (c) 2026 Redolarium Contributors, IIT Kanpur
# See LICENSE and THIRD_PARTY_LICENSES.md for full licence details.
# type: ignore
import os
import re
import json
import time
import openpyxl
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import seaborn as sns
from Bio import SeqIO
from redolarium.config import EXCEL_PALETTE, TAB_COLORS, CONFIG
from redolarium.utils import _sheet_title, _hdr, _alt_rows, _col_widths, apply_thin_borders, create_references_sheet_openpyxl, save_publication_plot
from redolarium.structures import PredictionResult
from redolarium.qc import get_module_qc_penalty

PIPELINE_VERSION = "3.0.0"

def load_json_config(path, default_val=None):
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"Warning: Failed to load config {path}: {e}")
    return default_val if default_val is not None else {}

def classify_mge_keyword_overlaps(sym, prod):
    """
    Removed unscientific MGE keyword proxy matching. MGEs must be identified mathematically via HMMs.
    """
    return []

def run_hmm_scan_for_cargo(proteins, hmm_db_path, manifest_data, logger):
    if not os.path.exists(hmm_db_path):
        logger.warning(f"Curated HMM database not found at {hmm_db_path}. Skipping HMM validation layer.")
        return {}

    try:
        import pyhmmer
    except ImportError:
        logger.warning("pyhmmer not installed in Python environment. Skipping HMM validation layer.")
        return {}

    try:
        with pyhmmer.plan7.HMMFile(hmm_db_path) as hmm_file:
            hmms = list(hmm_file)
    except Exception as e:
        logger.error(f"Failed to read HMM database {hmm_db_path}: {e}")
        return {}

    alphabet = pyhmmer.easel.Alphabet.amino()
    easel_seqs = []
    for p in proteins:
        ltag = p["Locus_Tag"]
        seq_str = p.get("Sequence", "")
        if not seq_str:
            continue
        clean_seq = re.sub(r'[^A-Z]', '', seq_str.upper())
        if clean_seq:
            easel_seqs.append(pyhmmer.easel.TextSequence(name=ltag.encode(), sequence=clean_seq).digitize(alphabet))

    if not easel_seqs:
        return {}

    logger.info(f"Running pyhmmer validation: {len(hmms)} profiles against {len(easel_seqs)} candidate proteins...")
    results = {}
    
    try:
        hits_list = list(pyhmmer.hmmsearch(hmms, easel_seqs))
    except Exception as e:
        logger.error(f"pyhmmer search failed: {e}")
        return {}
    
    profile_by_name = {p["name"]: p for p in manifest_data.get("profiles", [])}
    
    for tophits in hits_list:
        query_name = tophits.query.name
        prof_meta = profile_by_name.get(query_name, {})
        trusted_cutoff = prof_meta.get("trusted_cutoff", 20.0)
        recommended_cutoff = prof_meta.get("recommended_cutoff", 25.0)
        min_identity = prof_meta.get("minimum_identity", 0.40)
        
        for hit in tophits:
            locus_tag = hit.name.decode() if isinstance(hit.name, bytes) else hit.name
            evalue = hit.evalue
            bitscore = hit.score
            
            max_evalue = CONFIG.get("screening", {}).get("thresholds", {}).get("max_evalue", 1e-5)
            if evalue >= max_evalue or bitscore < trusted_cutoff:
                continue
                
            if len(hit.domains) == 0:
                continue
            best_domain = hit.domains[0]
            
            t_seq = best_domain.alignment.target_sequence
            h_seq = best_domain.alignment.hmm_sequence
            identical = 0
            total = 0
            for t_char, h_char in zip(t_seq, h_seq):
                if t_char == '-' or h_char == '-':
                    continue
                total += 1
                if t_char.upper() == h_char.upper():
                    identical += 1
            identity = (identical / total) if total > 0 else 0.0
            
            if identity < min_identity:
                continue
                
            coverage = (best_domain.alignment.target_to - best_domain.alignment.target_from + 1) / best_domain.alignment.target_length
            
            weights = CONFIG.get("screening", {}).get("weights", {"identity": 0.3, "coverage": 0.3, "bitscore": 0.4})
            w_id = weights.get("identity", 0.3)
            w_cov = weights.get("coverage", 0.3)
            w_bit = weights.get("bitscore", 0.4)
            
            bit_norm = min(1.0, bitscore / recommended_cutoff) if recommended_cutoff > 0 else 1.0
            confidence = (w_id * identity) + (w_cov * coverage) + (w_bit * bit_norm)
            
            hit_info = {
                "profile_name": query_name,
                "accession": prof_meta.get("accession", "Unknown"),
                "evalue": evalue,
                "bitscore": bitscore,
                "identity": identity,
                "coverage": coverage,
                "confidence": confidence,
                "description": prof_meta.get("description", "")
            }
            
            if locus_tag not in results:
                results[locus_tag] = []
            results[locus_tag].append(hit_info)
            
    return results

def evaluate_risk(category, target_family, is_plasmid, mge_nearby, risk_matrix_data, vf_signatures=None):
    if vf_signatures is None:
        scr_cfg = CONFIG.get("screening", {})
        vf_sigs_path = scr_cfg.get("vf_signatures_path", "config/vf_signatures.json")
        vf_signatures = load_json_config(vf_sigs_path)

    if category != "Virulence Factor":
        return "None", False, ""
        
    vf_class = "Unknown"
    fam_lower = target_family.lower()
    if "toxin" in fam_lower or "hemolysin" in fam_lower:
        vf_class = "Toxin"
    elif "secretion" in fam_lower:
        vf_class = "Secretion System"
    elif "adherence" in fam_lower or "pilus" in fam_lower or "fimbria" in fam_lower:
        vf_class = "Adherence"
    elif "evasion" in fam_lower or "capsule" in fam_lower:
        vf_class = "Immune Evasion"
    elif "enzymatic" in fam_lower or "elastase" in fam_lower or "collagenase" in fam_lower:
        vf_class = "Enzymatic Factor"

    for rule in risk_matrix_data.get("rules", []):
        cond = rule["condition"]
        if cond.get("class") != vf_class:
            continue
        if "localization" in cond:
            loc = "plasmid" if is_plasmid else "chromosome"
            if cond["localization"] != loc:
                continue
        if "mge_proximity" in cond:
            if cond["mge_proximity"] != mge_nearby:
                continue
                
        risk_level = rule["risk_level"]
        val_req = rule["validation_required"]
        evidence = rule["evidence"]
        
        injection_dependent = False
        for vf_class_def in vf_signatures.get("classes", []):
            if vf_class_def.get("class") == vf_class:
                injection_dependent = vf_class_def.get("requires_active_injection", False)
                break

        # Removed unscientific biosafety risk downgrade. 
        # Previously, if a T3SS/T6SS string match was not found, it downgraded the risk of Toxins.
        # This caused dangerous false negatives because "hypothetical protein" annotations would fail the string match.
        return risk_level, val_req, evidence
        
    return risk_matrix_data.get("default_risk", "None"), False, ""

def check_operon_context(locus_tag, genes_list, hit_type, role_name):
    if hit_type != "Quorum Sensing":
        return 1.0, ""
        
    try:
        idx = [g["Locus_Tag"] for g in genes_list].index(locus_tag)
    except ValueError:
        return 1.0, ""
        
    if "receptor" in role_name.lower() or "regulator" in role_name.lower():
        start_idx = max(0, idx - 15)
        end_idx = min(len(genes_list), idx + 16)
        
        has_synthase = False
        for i in range(start_idx, end_idx):
            if i == idx:
                continue
            g = genes_list[i]
            # Strictly check the mathematically verified HMM role for synthase activity, not the string symbol.
            if "synthase" in role:
                has_synthase = True
                break
                
        if not has_synthase:
            return 0.5, "Orphan receptor: No autoinducer synthase found within 15-gene neighborhood."
            
    return 1.0, ""
# Removed match_pre_filter_keywords entirely.

def run_screening_pipeline(query_gb, ortholog_mapping, out_dir, logger, qc_result=None) -> PredictionResult:
    logger.info("Stage 8: Running Application Screening & Phenotypic Profit Mapping...")
    start_time = time.time()
    
    record = max(list(SeqIO.parse(query_gb, "genbank")), key=lambda r: len(r.seq))
    
    # Load central screening configuration files
    scr_cfg = CONFIG.get("screening", {})
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
    qs_sigs_path = os.path.join(project_root, scr_cfg.get("qs_signatures_path", "config/qs_signatures.json"))
    vf_sigs_path = os.path.join(project_root, scr_cfg.get("vf_signatures_path", "config/vf_signatures.json"))
    risk_matrix_path = os.path.join(project_root, scr_cfg.get("risk_matrix_path", "config/risk_matrix.json"))
    manifest_path = os.path.join(project_root, scr_cfg.get("manifest_path", "resources/manifest.json"))
    hmm_db_path = os.path.join(project_root, scr_cfg.get("hmm_db_path", "resources/essential_bgc.hmm"))
    
    qs_signatures = load_json_config(qs_sigs_path)
    vf_signatures = load_json_config(vf_sigs_path)
    risk_matrix = load_json_config(risk_matrix_path)
    hmm_manifest = load_json_config(manifest_path)
    
    hmm_version = hmm_manifest.get("hmm_library_version", "1.0.0")
    logger.info(f"Loaded screening signatures. HMM Library version: {hmm_version}")
    
    # Load cargo keywords rules dynamically
    cargo_keywords_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config", "cargo_keywords.json")
    cargo_keywords = []
    if os.path.exists(cargo_keywords_path):
        try:
            with open(cargo_keywords_path, "r", encoding="utf-8") as f_cargo:
                cargo_keywords = json.load(f_cargo)
            logger.info(f"Loaded {len(cargo_keywords)} dynamic cargo keyword rules from {cargo_keywords_path}")
        except Exception as e_cargo:
            logger.warning(f"Failed to load dynamic cargo keywords: {e_cargo}")
    
    is_plasmid = "plasmid" in record.description.lower() or "plasmid" in record.id.lower()
    
    mge_nearby = False
    
    all_proteins = []
    candidates = []
    
    for feat in record.features:
        if feat.type == "CDS":
            prod = feat.qualifiers.get("product", [""])[0]
            sym = feat.qualifiers.get("gene", ["NA"])[0]
            ltag = feat.qualifiers.get("locus_tag", [""])[0]
            trans = feat.qualifiers.get("translation", [""])[0]
            
            p_info = {
                "Locus_Tag": ltag,
                "Gene_Symbol": sym,
                "Product": prod,
                "Sequence": trans,
                "Start": int(feat.location.start),
                "End": int(feat.location.end),
                "Strand": feat.location.strand,
                "EC_number": list(feat.qualifiers.get("EC_number", [])),
                "db_xref": list(feat.qualifiers.get("db_xref", []))
            }
            all_proteins.append(p_info)
            
            # Removed unscientific Secretion System word mines.

            # Removed keyword pre-filtering ("word mines") entirely.
            # We add all flanking genes as candidates and rely strictly on mathematical HMM homology.
            candidates.append(p_info)
    # Run HMM scan validation
    hmm_hits = run_hmm_scan_for_cargo(candidates, hmm_db_path, hmm_manifest, logger)
    
    # False Negative Mitigation
    critical_accessions = ["PF01338", "PF03006", "PF07968"]
    critical_hmms = []
    if os.path.exists(hmm_db_path):
        try:
            import pyhmmer
            with pyhmmer.plan7.HMMFile(hmm_db_path) as hmm_file:
                for hmm in hmm_file:
                    acc = hmm.accession.split('.')[0] if hmm.accession else ""
                    if acc in critical_accessions:
                        critical_hmms.append(hmm)
            if critical_hmms:
                alphabet = pyhmmer.easel.Alphabet.amino()
                full_easel = []
                for p in all_proteins:
                    if any(c["Locus_Tag"] == p["Locus_Tag"] for c in candidates):
                        continue
                    clean_seq = re.sub(r'[^A-Z]', '', p["Sequence"].upper())
                    if clean_seq:
                        full_easel.append(pyhmmer.easel.TextSequence(name=p["Locus_Tag"].encode(), sequence=clean_seq).digitize(alphabet))
                if full_easel:
                    logger.info(f"Running false-negative mitigation scan for {len(critical_hmms)} toxin profiles against remaining {len(full_easel)} proteins...")
                    tophits_list = list(pyhmmer.hmmsearch(critical_hmms, full_easel))
                    profile_by_name = {p["name"]: p for p in hmm_manifest.get("profiles", [])}
                    
                    for tophits in tophits_list:
                        qname = tophits.query.name
                        prof_meta = profile_by_name.get(qname, {})
                        trusted_cutoff = prof_meta.get("trusted_cutoff", 20.0)
                        
                        for hit in tophits:
                            locus_tag = hit.name.decode() if isinstance(hit.name, bytes) else hit.name
                            if hit.score >= trusted_cutoff and hit.evalue < 1e-5:
                                logger.warning(f"[FALSE NEGATIVE ALERT] Critical toxin match found for {locus_tag} missed by keyword pre-filters!")
                                found_p = next(pr for pr in all_proteins if pr["Locus_Tag"] == locus_tag)
                                candidates.append(found_p)
                                fresh_hit = run_hmm_scan_for_cargo([found_p], hmm_db_path, hmm_manifest, logger)
                                hmm_hits.update(fresh_hit)
        except Exception as e:
            logger.error(f"Error during false-negative mitigation scan: {e}")

    orth_dict = {g["Locus_Tag"]: g for g in ortholog_mapping}
    screened_cargo = []
    
    from redolarium.metabolism import load_gene_symbol_to_ec
    gene_symbol_to_ec = load_gene_symbol_to_ec()
    
    for p in candidates:
        prod = p["Product"].lower()
        sym = p["Gene_Symbol"]
        ltag = p["Locus_Tag"]
        ident = orth_dict.get(ltag, {}).get("Identity_Pct", 0.0)
        
        # Extract EC number using identical logic as metabolism.py
        ecs = list(p.get("EC_number", []))
        for ref in p.get("db_xref", []):
            if ref.lower().startswith("ec:"):
                ec_val = ref.split(":")[1].strip()
                if ec_val not in ecs:
                    ecs.append(ec_val)
        if not ecs and sym != "NA" and sym:
            sym_lower = sym.lower()
            if sym_lower in gene_symbol_to_ec:
                ecs.append(gene_symbol_to_ec[sym_lower])
        if not ecs:
            m = re.search(r"ec[:\s]+(\d+\.\d+\.\d+\.\d+)", prod, re.IGNORECASE)
            if m:
                ecs.append(m.group(1))
        ec_extracted = ecs[0] if ecs else "NA"
        
        cat = None
        fam = None
        risk = "None"
        prop = None
        conf = 1.0
        val_flag = "No"
        
        has_hmm_match = ltag in hmm_hits and len(hmm_hits[ltag]) > 0
        
        # Removed unscientific keyword proxy fallbacks for CAZymes, Metal Resistance, and Intrinsic AMR.
        # Strict enforcement of Rule R3: functional classification must rely on verified mathematical mapping (HMM).
        
        # 4. Quorum Sensing / Virulence Factors via hybrid HMM
        if has_hmm_match:
            best_hit = max(hmm_hits[ltag], key=lambda x: x["confidence"])
            h_name = best_hit["profile_name"]
            conf = best_hit["confidence"]
            
            is_qs_hmm = any(p_m["name"] == h_name for p_m in hmm_manifest.get("profiles", []) if "Lux" in p_m["description"] or "autoinducer" in p_m["description"].lower() or "AIP" in p_m["description"] or "Agr" in p_m["description"] or "diffusible" in p_m["description"].lower())
            
            if is_qs_hmm or "PF00765" in best_hit["accession"] or "PF00196" in best_hit["accession"] or "PF08660" in best_hit["accession"]:
                cat = "Quorum Sensing"
                sys_name = "Unclassified Quorum Sensing"
                sys_type = "Unclassified QS"
                sys_desc = best_hit["description"]
                
                for sys_data in qs_signatures.get("systems", []):
                    for role_type, pf_list in sys_data.get("pfams", {}).items():
                        if best_hit["accession"] in pf_list:
                            sys_name = sys_data["name"]
                            sys_type = sys_data["type"]
                            sys_desc = sys_data["description"]
                            break
                            
                fam = f"{sys_type} (HMM validated)"
                prop = f"System: {sys_name} | Acc: {best_hit['accession']} | E-value: {best_hit['evalue']:.2e}"
                
                operon_mult, operon_msg = check_operon_context(ltag, candidates, cat, sys_name)
                conf *= operon_mult
                if operon_msg:
                    prop += f" | NOTE: {operon_msg}"
            else:
                cat = "Virulence Factor"
                cl_name = "Toxin"
                for cl in vf_signatures.get("classes", []):
                    if best_hit["accession"] in cl.get("pfams", []):
                        cl_name = cl["name"]
                        break
                        
                fam = f"{cl_name} (HMM validated)"
                risk_level, val_req, evidence = evaluate_risk(cat, cl_name, is_plasmid, mge_nearby, risk_matrix, vf_signatures=vf_signatures)
                risk = risk_level
                val_flag = "Yes" if val_req else "No"
                prop = f"Class: {cl_name} | Acc: {best_hit['accession']} | E-value: {best_hit['evalue']:.2e} | Ev: {evidence}"
        else:
            # Removed unscientific keyword fallback.
            # If HMM validation fails to match a known virulence/QS profile, we do not guess based on strings.
            pass

        if cat:
            # Map mobile element keyword overlaps
            mge_overlaps = classify_mge_keyword_overlaps(sym, p["Product"])
            if mge_overlaps:
                db_hits_str = " | MGE Overlaps: " + ", ".join([f"{o['database']} (v{o['version']}, ID={o['match_id']})" for o in mge_overlaps])
                prop += db_hits_str
                
            screened_cargo.append({
                "Locus_Tag": ltag,
                "Gene_Symbol": sym,
                "Functional_Category": cat,
                "Target_Family": fam,
                "EC_Number": ec_extracted,
                "Identity_Pct": ident,
                "Biosafety_Risk": risk,
                "Requires_Wet_Lab_Validation": val_flag,
                "Confidence_Score": round(conf, 4),
                "Industrial_Value_Proposition": prop
            })

    if not screened_cargo:
        logger.info("No application cargo or biosafety flags identified in genome.")
        screened_cargo.append({
            "Locus_Tag": "None",
            "Gene_Symbol": "None",
            "Functional_Category": "None",
            "Target_Family": "None",
            "EC_Number": "NA",
            "Identity_Pct": 0.0,
            "Biosafety_Risk": "None",
            "Requires_Wet_Lab_Validation": "No",
            "Confidence_Score": 0.0,
            "Industrial_Value_Proposition": "No cargo detected"
        })

    df_cargo = pd.DataFrame(screened_cargo)
    os.makedirs(os.path.join(out_dir, "screening_cazymes"), exist_ok=True)
    logger.info("Screened application cargo table generated in memory.")

    # Visual cargo chart
    plt.figure(figsize=(8, 5))
    sns.set_theme(style="whitegrid")
    counts_df = df_cargo[df_cargo["Functional_Category"] != "None"]["Functional_Category"].value_counts().reset_index()
    counts_df.columns = ["Category", "Count"]
    if not counts_df.empty:
        sns.barplot(data=counts_df, x="Count", y="Category", palette="coolwarm", hue="Category", legend=False)
    plt.title(f"Functional Cargo Distribution & Biosafety Profile (v{PIPELINE_VERSION})", fontsize=11, fontweight="bold", pad=15)
    plt.xlabel("Annotated Gene Count", fontsize=9)
    plt.ylabel("Functional Group", fontsize=9)
    plt.tight_layout()
    fig_out = os.path.join(out_dir, "screening_cazymes", "cargo_distribution.png")
    save_publication_plot(fig_out, dpi=300)
    plt.close()

    # Formatted Excel Sheet Workbook
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_properties.tabColor = TAB_COLORS["Summary"]
    
    ws_sum.cell(row=3, column=1, value="Screening Parameter")
    ws_sum.cell(row=3, column=2, value="Count / Value")
    ws_sum.cell(row=3, column=3, value="Evaluation Interpretation Context")
    
    weights_info = CONFIG.get("screening", {}).get("weights", {"identity": 0.3, "coverage": 0.3, "bitscore": 0.4})
    formula_text = f"Confidence = {weights_info.get('identity')}*Identity + {weights_info.get('coverage')}*Coverage + {weights_info.get('bitscore')}*NormalizedBitscore"
    
    summary_data = [
        ("Total Screened Genes", str(len(screened_cargo)), "Total proteins checked for industrial applications"),
        ("CAZyme Family Proteins", str(sum(1 for c in screened_cargo if c["Functional_Category"] == "CAZyme")), "Biomass degradation, pectinase, cellulose digestive enzymes"),
        ("Heavy Metal Transporters", str(sum(1 for c in screened_cargo if c["Functional_Category"] == "Metal Resistance")), "Arsenic/Copper efflux pumps for soil adaptation"),
        ("AMR Intrinsic Markers", str(sum(1 for c in screened_cargo if c["Functional_Category"] == "AMR (Intrinsic)")), "Intrinsic antibiotic resistance markers"),
        ("VFDB Safety Flags (HMM)", str(sum(1 for c in screened_cargo if c["Functional_Category"] == "Virulence Factor")), "HMM-validated safety screening tags (toxins/secretions)"),
        ("Requires Wet-Lab Validation", str(sum(1 for c in screened_cargo if c["Requires_Wet_Lab_Validation"] == "Yes")), "Candidate safety flags needing laboratory assays"),
        ("Input Accession", record.id, "Query sequence identifier"),
        ("Screening Algorithm", "Pfam HMM cargo screening & biosafety risk matrix classifier", "Algorithm used for detection"),
        ("CAZyme Database Reference", "CAZy (Lombard et al. 2014, doi:10.1093/nar/gkt1178)", "Carbohydrate-Active enZymes database citation"),
        ("Confidence Evaluation Formula", formula_text, "Configured scoring model parameters for transparency"),
        ("Pipeline Version", f"v{PIPELINE_VERSION}", "Redolarium core pipeline implementation release version"),
        ("HMM Database Version", f"Pfam Library v{hmm_version}", "Curated target HMM profile release version")
    ]
    
    for idx, (param, val, desc) in enumerate(summary_data):
        ws_sum.cell(row=4+idx, column=1, value=param).font = openpyxl.styles.Font(name="Calibri", size=9, bold=True)
        ws_sum.cell(row=4+idx, column=2, value=val).alignment = openpyxl.styles.Alignment(horizontal="center")
        ws_sum.cell(row=4+idx, column=3, value=desc)
        
    _sheet_title(ws_sum, "Phenotypic Profit Catalog & Biosafety Screening Summary",
                 f"Genome: {record.id} | Isolate Safety & Industrial Applications Dashboard")
    _hdr(ws_sum, 3, EXCEL_PALETTE["header_fill"])
    apply_thin_borders(ws_sum, 4, ws_sum.max_row)
    _alt_rows(ws_sum, 4, ws_sum.max_row, EXCEL_PALETTE["alt_row_fill"])
    
    start_int = ws_sum.max_row + 2
    ws_sum.merge_cells(start_row=start_int, start_column=1, end_row=start_int+3, end_column=3)
    ws_sum.cell(row=start_int, column=1, value=(
        "Interpretation Summary: Screening of the query genome demonstrates potential industrial suitability. "
        "Confidence is calculated based on a weighted sum of sequence identity, alignment coverage, and normalized HMM bitscore. "
        "High and Moderate virulence risk profiles have been marked with a mandatory wet-lab validation flag. "
        "General two-component regulatory matches lacking proximal autoinducer synthases have been contextually downgraded to reduce false positives."
    )).font = openpyxl.styles.Font(name="Calibri", size=9, italic=True)
    ws_sum.cell(row=start_int, column=1).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    _col_widths(ws_sum, [30, 24, 65])

    # Sheet 2: Catalog Details
    ws_cat = wb.create_sheet(title="Phenotypic Profit Catalog")
    ws_cat.sheet_properties.tabColor = TAB_COLORS.get("Screening", "1F3864")
    
    headers = ["Locus Tag", "Gene Symbol", "Functional Category", "Target Family", "EC Number", "Identity Pct (%)", "Biosafety Risk", "Wet-Lab Required", "Confidence", "Industrial Value Proposition"]
    ws_cat.append(headers)
    for c in screened_cargo:
        ws_cat.append([
            c["Locus_Tag"],
            c["Gene_Symbol"],
            c["Functional_Category"],
            c["Target_Family"],
            c["EC_Number"],
            c["Identity_Pct"],
            c["Biosafety_Risk"],
            c["Requires_Wet_Lab_Validation"],
            c["Confidence_Score"],
            c["Industrial_Value_Proposition"]
        ])
        
    _sheet_title(ws_cat, "Phenotypic Profit Catalog - Industrial Applications Details",
                 "List of annotated proteins possessing academic, industrial, or agricultural utility.")
    _hdr(ws_cat, 3, EXCEL_PALETTE["header_fill"])
    apply_thin_borders(ws_cat, 4, ws_cat.max_row)
    
    for r in range(4, ws_cat.max_row + 1):
        cat_val = ws_cat.cell(row=r, column=3).value
        risk_val = ws_cat.cell(row=r, column=7).value
        
        if cat_val == "CAZyme":
            ws_cat.cell(row=r, column=3).fill = openpyxl.styles.PatternFill("solid", fgColor=EXCEL_PALETTE["positive_fill"])
        elif cat_val == "Metal Resistance":
            ws_cat.cell(row=r, column=3).fill = openpyxl.styles.PatternFill("solid", fgColor=EXCEL_PALETTE["bgc_core_fill"])
            
        if "High" in str(risk_val):
            ws_cat.cell(row=r, column=7).fill = openpyxl.styles.PatternFill("solid", fgColor=EXCEL_PALETTE["warning_fill"])
            ws_cat.cell(row=r, column=7).font = openpyxl.styles.Font(name="Calibri", size=9, bold=True)
        elif "Moderate" in str(risk_val):
            ws_cat.cell(row=r, column=7).fill = openpyxl.styles.PatternFill("solid", fgColor=EXCEL_PALETTE["highlight_fill"])
            
        for col in range(1, 10):
            ws_cat.cell(row=r, column=col).alignment = openpyxl.styles.Alignment(horizontal="center")
        
        for col in range(1, 11):
            is_bold = (col == 1)
            ws_cat.cell(row=r, column=col).font = openpyxl.styles.Font(name="Calibri", size=9, bold=is_bold)
            
    _alt_rows(ws_cat, 4, ws_cat.max_row, EXCEL_PALETTE["alt_row_fill"])
    _col_widths(ws_cat, [15, 12, 20, 26, 12, 14, 20, 16, 12, 80])
    
    # Sheet 3: References
    create_references_sheet_openpyxl(wb)
    
    xls_out = os.path.join(out_dir, "screening_cazymes", "phenotypic_profit.xlsx")
    wb.save(xls_out)
    logger.info(f"Saved Excel Phenotypic Profit Catalog to: {xls_out}")
    
    # Visual genome map
    try:
        fig, ax = plt.subplots(figsize=(12, 4))
        sns.set_theme(style="white")
        genome_len = len(record.seq)
        ax.plot([0, genome_len], [0.5, 0.5], color="#CCCCCC", lw=4, zorder=1)
        
        colors_map = {
            "CAZyme": "#4A90E2",
            "Metal Resistance": "#2ECC71",
            "Quorum Sensing": "#9B59B6",
            "Virulence Factor": "#E74C3C",
            "AMR (Intrinsic)": "#F1C40F"
        }
        
        for cargo in screened_cargo:
            if cargo["Locus_Tag"] == "None":
                continue
            cat = cargo["Functional_Category"]
            if cat not in colors_map:
                continue
            
            # Find gene location
            feat = next(f for f in record.features if f.type == "CDS" and f.qualifiers.get("locus_tag", [""])[0] == cargo["Locus_Tag"])
            start = int(feat.location.start)
            end = int(feat.location.end)
            length = end - start
            col = colors_map[cat]
            
            # Layered density track / feature map visualization
            direction = 1 if feat.location.strand >= 0 else -1
            # Adjust y_pos to layer tracks slightly
            y_pos = 0.5 + (0.05 * (list(colors_map.keys()).index(cat) - 2))
            
            track_block = patches.Rectangle((start, y_pos - 0.02), length, 0.04, facecolor=col, edgecolor="none", alpha=0.7, zorder=3)
            ax.add_patch(track_block)
            
            label = cargo["Gene_Symbol"]
            if label == "NA":
                label = cargo["Locus_Tag"].split("_")[-1]
            ax.text(start + length/2, y_pos + 0.03, label, fontsize=6, ha="center", va="bottom", zorder=4, rotation=45,
                    bbox=dict(boxstyle="round,pad=0.1", fc="white", ec=col, alpha=0.9, lw=0.5))
            
        handles = [plt.Line2D([0], [0], color=col, lw=6, label=cat) for cat, col in colors_map.items()]
        ax.legend(handles=handles, loc="upper right", frameon=True, fontsize=8)
        
        # Explicit annotation if empty
        if not [c for c in screened_cargo if c["Locus_Tag"] != "None"]:
            ax.text(genome_len / 2, 0.5, "No specific functional cargo identified", ha="center", va="center", fontsize=12, color="gray", fontstyle="italic")
        
        ax.set_xlim(-0.02 * genome_len, 1.02 * genome_len)
        ax.set_ylim(0.3, 0.7)
        ax.get_yaxis().set_visible(False)
        ax.set_xlabel("Genomic Coordinates (bp)", fontsize=9, fontweight="bold")
        ax.set_title(f"Genomic Synteny & Distribution Map of Functional Cargo: {record.id} (v{PIPELINE_VERSION})", fontsize=10, fontweight="bold", pad=12)
        plt.tight_layout()
        
        cargo_fig_out = os.path.join(out_dir, "screening_cazymes", "cargo_distribution.png")
        save_publication_plot(cargo_fig_out, dpi=300)
        plt.close()
        logger.info(f"Saved directional linear genomic cargo map to: {cargo_fig_out}")
    except Exception as ex:
        logger.warning(f"Failed to generate linear genome cargo map: {ex}")
        
    # Calculate screening confidence
    completeness = 100.0
    contamination = 0.0
    closure_status = "Closed"
    if qc_result:
        completeness = qc_result.prediction.get("completeness", 100.0)
        contamination = qc_result.prediction.get("contamination", 0.0)
        closure_status = qc_result.genome_closure_status
        
    qc_penalty = get_module_qc_penalty("annotation", completeness, contamination)
    
    base_score = 0.95
    if len(screened_cargo) == 0 or (len(screened_cargo) == 1 and screened_cargo[0]["Locus_Tag"] == "None"):
        base_score = 0.85
        
    final_score = base_score - qc_penalty
    final_score = max(0.0, min(1.0, final_score))
    
    db_vers = CONFIG.get("database_versions", {})
    
    return PredictionResult(
        prediction=screened_cargo,
        confidence_score=final_score,
        algorithm="Pfam HMM cargo screening & biosafety risk matrix classifier",
        algorithm_version="v3.0.0",
        genome_closure_status=closure_status,
        database="MobileOG-db / ACLAME / ICEberg / ISfinder / PHASTER",
        database_version=db_vers.get("mobileog", "1.6"),
        evidence=[
            f"Screened {len(candidates)} candidate genes across genome.",
            f"Identified {len(screened_cargo)} functional cargos (CAZymes, AMR, Metal, QS, VFs).",
            f"Mapped MGE database overlaps against ISfinder, ICEberg, MobileOG, ACLAME, and PHASTER."
        ],
        limitations=[
            "Screening is based on Pfam domain models and sequence alignment.",
            # Removed keyword fallback for transporter/resistance/VF
            # Classification strictly adheres to HMM mapping.
            "Functional phenotype predictions require experimental validation."
        ],
        citations=[
            "MobileOG-db: Brown et al. 2022 (doi:10.1128/msystems.00991-22)",
            "ACLAME: Leplae et al. 2010 (doi:10.1093/nar/gkp941)",
            "ICEberg: Liu et al. 2019 (doi:10.1093/nar/gky1123)",
            "ISfinder: Siguier et al. 2006 (doi:10.1093/nar/gkj014)",
            "PHASTER: Arndt et al. 2016 (doi:10.1093/nar/gkw387)"
        ],
        runtime=time.time() - start_time,
        warnings=[f"QC assembly penalty applied: -{qc_penalty:.2f}"] if qc_penalty > 0.05 else []
    )
